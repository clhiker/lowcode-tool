#!/usr/bin/env python3
"""
多模型并发分类器：将模板按批次放入任务池，多模型并发消费。

流程：
  1. 逐个处理 module-list/ 下的 xlsx 文件
  2. 对每个文件，将所有行按 batch_size 切分为批次，放入队列
  3. 启动 N 个 worker 线程（每个使用不同的模型配置），并发从队列取任务
  4. 每个 worker 独立调用 API、解析、校验、重试
  5. 所有批次完成后，统一写回 xlsx

用法：
  python3 classifier/batch_run.py
"""

import configparser
import queue
import re
import sys
import time
import argparse
import threading
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openai import OpenAI

from prompts import SYSTEM_PROMPT, build_batch_prompt


# ==================== 工具函数 ====================

def load_gbt_codes(gbt_path):
    codes = set()
    with open(gbt_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            m = re.match(r"(\d{4})\s+", line)
            if m:
                codes.add(m.group(1))
    print(f"已加载 {len(codes)} 个有效 GB/T 4754-2017 小类代码", file=sys.stderr)
    return codes


def extract_codes_from_domain_str(domain_str):
    return re.findall(r"\b(\d{4})\b", domain_str)


def validate_domains(domains_list, valid_codes):
    invalid_items = []
    for i, ds in enumerate(domains_list):
        codes = extract_codes_from_domain_str(ds)
        bad = [c for c in codes if c not in valid_codes]
        if bad:
            invalid_items.append((i, bad))
    return invalid_items


def parse_batch_results(text, n):
    domains_list = [""] * n
    types_list = [""] * n
    current_idx = -1
    for line in text.strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"<(\d+)>\s*$", line)
        if m:
            current_idx = int(m.group(1)) - 1
            continue
        dm = re.match(r"领域[：:]\s*(.*)", line)
        if dm and 0 <= current_idx < n:
            domains_list[current_idx] = dm.group(1).strip()
            continue
        tm = re.match(r"类型[：:]\s*(.*)", line)
        if tm and 0 <= current_idx < n:
            types_list[current_idx] = tm.group(1).strip()
    for i in range(n):
        if not domains_list[i]:
            domains_list[i] = "通用（前端表单）"
        if not types_list[i]:
            types_list[i] = "组件模板"
    return domains_list, types_list


def call_api(client, model, messages, max_tokens=8192, timeout=180):
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=messages,
            temperature=0.1,
            max_tokens=max_tokens,
            timeout=timeout,
        )
        choice = resp.choices[0]
        content = choice.message.content
        if content is None:
            finish = choice.finish_reason
            print(f"  [警告] content为空, finish_reason={finish}", file=sys.stderr)
            return None
        return content.strip()
    except Exception as e:
        print(f"  [错误] API调用失败: {e}", file=sys.stderr)
        return None


def evaluate_formula(val, row_idx):
    s = str(val)
    m = re.search(
        r'CONCAT\("([^"]+)"\s*,\s*TEXT\s*\(\s*ROW\(\)\s*-\s*(\d+)\s*,\s*"([^"]+)"\s*\)\s*\)',
        s,
    )
    if m:
        prefix = m.group(1)
        offset = int(m.group(2))
        fmt = m.group(3)
        num = row_idx - offset
        return f'{prefix}{num:0{len(fmt)}}'
    return s


def load_settings(config_path):
    config = configparser.ConfigParser()
    config.read(config_path, encoding="utf-8")
    s = config["settings"] if config.has_section("settings") else {}
    return {
        "gbt": s.get("gbt", "classifier/gbt4754_2017_小类.txt"),
        "batch_size": int(s.get("batch_size", 50)),
        "retry_delay": int(s.get("retry_delay", 5)),
        "models": s.get("models", ""),
    }


def load_model_configs(config_path, selected_names=None):
    config = configparser.ConfigParser()
    config.read(config_path, encoding="utf-8")
    all_sections = [s for s in config.sections() if s != "settings"]

    if selected_names:
        names = [n.strip() for n in selected_names.split(",") if n.strip()]
        sections = [s for s in all_sections if s in names]
        if not sections:
            missing = set(names) - set(all_sections)
            print(f"错误：settings.models 中指定的以下模型不存在: {missing}", file=sys.stderr)
            sys.exit(1)
    else:
        sections = all_sections

    models = []
    for section in sections:
        models.append({
            "name": section,
            "base_url": config[section]["base_url"].rstrip("/"),
            "api_key": config[section]["api_key"],
            "model": config[section]["model"],
            "max_tokens": int(config[section].get("max_tokens", 4096)),
        })
    if not models:
        print("错误：config.ini 中未找到任何模型配置", file=sys.stderr)
        sys.exit(1)
    print(f"发现 {len(models)} 个模型配置:")
    for m in models:
        print(f"  [{m['name']}] {m['model']}")
    return models


# ==================== 任务 & 线程安全写入 ====================

@dataclass
class BatchTask:
    batch_id: int
    items: list
    indices: list


class SafeXlsxWriter:
    """线程安全地逐批写入 worksheet 和临时文件，全部完成后一次性保存 xlsx。

    tmp 文件格式（每行）: {全局索引}\\t{领域}\\t{类型}

    支持断点续传：启动时从 tmp 恢复已完成条目，跳过对应批次。
    """

    def __init__(self, wb, ws, col_new1, col_new2, items, xlsx_path):
        self.wb = wb
        self.ws = ws
        self.col_new1 = col_new1
        self.col_new2 = col_new2
        self.items = items  # [(row_idx, tid, desc, func), ...]
        self.xlsx_path = xlsx_path
        self.tmp_path = xlsx_path.with_suffix(".tmp")
        self._lock = threading.Lock()
        self._completed = 0
        self._completed_lock = threading.Lock()
        self.print_lock = threading.Lock()
        self.total = len(items)

    def load_checkpoint(self):
        """读取 tmp 返回 {global_idx: (domain, type)}，不存在返回空 dict。"""
        if not self.tmp_path.exists():
            return {}
        result = {}
        with open(self.tmp_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                parts = line.split("\t")
                if len(parts) >= 3:
                    result[int(parts[0])] = (parts[1], parts[2])
        return result

    def write_batch(self, worker_id, batch_indices, domain_list, type_list):
        """写入 in-memory worksheet 并追加到临时文件。"""
        with self._lock:
            for k, idx in enumerate(batch_indices):
                row_idx = self.items[idx][0]
                self.ws.cell(row=row_idx, column=self.col_new1).value = domain_list[k]
                self.ws.cell(row=row_idx, column=self.col_new2).value = type_list[k]
            with open(self.tmp_path, "a", encoding="utf-8") as f:
                for k, idx in enumerate(batch_indices):
                    f.write(f"{idx}\t{domain_list[k]}\t{type_list[k]}\n")
        with self._completed_lock:
            self._completed += len(batch_indices)

    def save_and_cleanup(self):
        """保存 xlsx 并删除临时文件。"""
        with self._lock:
            self.wb.save(self.xlsx_path)
        if self.tmp_path.exists():
            self.tmp_path.unlink()

    @property
    def completed(self):
        with self._completed_lock:
            return self._completed

    def safe_print(self, model_name, batch_start, batch_size, *args, **kwargs):
        with self.print_lock:
            c = self.completed
            prefix = f"[{model_name}][{batch_start+1}-{batch_start+batch_size}/{self.total}]"
            print(prefix, *args, **kwargs)


# ==================== Worker 逻辑 ====================

def worker_loop(task_queue, writer, valid_codes, client, model_cfg, worker_id, retry_delay):
    while True:
        task = task_queue.get()
        if task is None:
            task_queue.task_done()
            break

        batch_items = [(d, f) for _, _, d, f in task.items]
        batch_size = len(batch_items)
        mtokens = max(8192, batch_size * 2048)

        retry_history = ""
        val_attempt = 0
        api_attempt = 0
        known_bad = {}

        while True:
            tag = f"[修正 {val_attempt}]" if val_attempt > 0 else ""
            api_tag = f"[重试 {api_attempt}]" if api_attempt > 0 else ""
            writer.safe_print(model_cfg["name"], task.indices[0], batch_size,
                              f"处理中{tag}{api_tag}...", end=" ", flush=True)

            user_prompt = build_batch_prompt(batch_items)
            if retry_history:
                user_prompt += "\n\n" + retry_history

            text = call_api(client, model_cfg["model"], [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ], max_tokens=mtokens)

            if not text:
                api_attempt += 1
                writer.safe_print(model_cfg["name"], task.indices[0], batch_size,
                                  f"→ API失败，{retry_delay}秒后重试")
                time.sleep(retry_delay)
                continue

            domains_list, types_list = parse_batch_results(text, batch_size)
            invalid = validate_domains(domains_list, valid_codes)

            if not invalid:
                writer.write_batch(worker_id, task.indices, domains_list, types_list)
                writer.safe_print(model_cfg["name"], task.indices[0], batch_size, "→ 通过")
                break

            # 累计本次的错误码
            for local_idx, bad_codes in invalid:
                if local_idx not in known_bad:
                    known_bad[local_idx] = set()
                known_bad[local_idx].update(bad_codes)

            val_attempt += 1

            if val_attempt >= 6:
                bad_set = {local_idx for local_idx, _ in invalid}
                bad_list = sorted(bad_set)
                patched = [(i, domains_list[i]) for i in bad_list]
                for local_idx in bad_list:
                    domains_list[local_idx] = "6513 应用软件开发, 6519 其他软件开发"
                writer.write_batch(worker_id, task.indices, domains_list, types_list)
                writer.safe_print(model_cfg["name"], task.indices[0], batch_size,
                                  f"→ 校验失败{val_attempt}次，第{len(bad_list)}项强制设为 6513 应用软件开发")
                break

            # 用累计的错误码构建修正提示
            bad_lines = []
            for local_idx in sorted(known_bad):
                codes = ", ".join(sorted(known_bad[local_idx]))
                bad_lines.append(f"  <{local_idx + 1}> 中曾出现 {codes}，请替换为真实存在的分类。")
            retry_history = (
                "注意：以下代码在之前的回答中反复使用了不存在的 GB/T 4754-2017 代码，请务必修正：\n"
                + "\n".join(bad_lines)
                + "\n只输出修正后的结果即可，格式与之前完全相同（每个模板一个块）。"
            )

        task_queue.task_done()


# ==================== 单文件处理 ====================

def process_xlsx(xlsx_path, model_configs, gbt_path, batch_size, retry_delay):
    print(f"\n{'='*60}")
    print(f"处理文件: {xlsx_path.name}")
    print(f"{'='*60}")

    valid_codes = load_gbt_codes(gbt_path)
    t0 = time.time()

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["汇总表"]

    col_id, col_func, col_desc = 1, 8, 11

    items = []
    for row_idx in range(2, ws.max_row + 1):
        tid_raw = ws.cell(row=row_idx, column=col_id).value
        desc = ws.cell(row=row_idx, column=col_desc).value
        if not tid_raw or not desc or not str(desc).strip():
            continue
        tid = evaluate_formula(tid_raw, row_idx)
        func = ws.cell(row=row_idx, column=col_func).value or ""
        items.append((row_idx, str(tid).strip(), str(desc).strip(), str(func).strip()))

    if not items:
        print("没有需要处理的模板。")
        return

    total = len(items)
    n_batches = (total + batch_size - 1) // batch_size
    print(f"共 {total} 个模板，batch-size={batch_size}，{n_batches} 个批次，{len(model_configs)} 个 worker 并发")

    # 找到/创建新增列
    col_new1 = col_new2 = None
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h == "新增列1":
            col_new1 = col
        elif h == "新增列2":
            col_new2 = col
    if col_new1 is None:
        col_new1 = ws.max_column + 1
        ws.cell(row=1, column=col_new1).value = "新增列1"
    if col_new2 is None:
        col_new2 = col_new1 + 1
        ws.cell(row=1, column=col_new2).value = "新增列2"
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col_new1).value = None
        ws.cell(row=row, column=col_new2).value = None

    # 创建线程安全写入器
    writer = SafeXlsxWriter(wb, ws, col_new1, col_new2, items, xlsx_path)

    # 断点续传：从 tmp 恢复已完成条目
    checkpoint = writer.load_checkpoint()
    done_indices = set()
    if checkpoint:
        done_indices = set(checkpoint.keys())
        print(f"发现断点文件，已有 {len(done_indices)}/{total} 个条目完成，跳过对应批次")
        for idx, (domain, type_) in checkpoint.items():
            row_idx = items[idx][0]
            ws.cell(row=row_idx, column=col_new1).value = domain
            ws.cell(row=row_idx, column=col_new2).value = type_
        writer._completed = len(done_indices)

    # 填充任务队列（跳过已完成的批次）
    task_queue = queue.Queue()
    queued = 0
    for start in range(0, total, batch_size):
        batch = items[start:start + batch_size]
        indices = list(range(start, start + len(batch)))
        if done_indices and done_indices.issuperset(set(indices)):
            continue
        task_queue.put(BatchTask(batch_id=start // batch_size, items=batch, indices=indices))
        queued += len(batch)

    if task_queue.empty():
        print("所有条目已完成，直接保存 xlsx")
        writer.save_and_cleanup()
        elapsed = time.time() - t0
        print(f"\n  ✓ 完成！耗时 {elapsed:.1f}s（已保存至 {xlsx_path.name}）")
        return

    print(f"剩余 {queued} 个模板，{task_queue.qsize()} 个批次，{len(model_configs)} 个 worker 并发")

    # 启动 worker 线程
    threads = []
    for i, mc in enumerate(model_configs):
        client = OpenAI(base_url=mc["base_url"], api_key=mc["api_key"])
        t = threading.Thread(
            target=worker_loop,
            args=(task_queue, writer, valid_codes, client, mc, i + 1, retry_delay),
            daemon=True,
        )
        t.start()
        threads.append(t)

    # 等待所有任务消费完毕
    task_queue.join()

    # 通知 workers 退出
    for _ in threads:
        task_queue.put(None)
    for t in threads:
        t.join()

    # 一次性保存 xlsx 并删除临时文件
    writer.save_and_cleanup()

    elapsed = time.time() - t0
    print(f"\n  ✓ 完成！耗时 {elapsed:.1f}s（已保存至 {xlsx_path.name}）")


# ==================== 入口 ====================

def main():
    parser = argparse.ArgumentParser(
        description="多模型并发模板分类"
    )
    parser.add_argument("--config", default="classifier/config.ini",
                        help="配置文件路径")
    parser.add_argument("--gbt", default=None,
                        help="GB/T 4754-2017 小类代码文件路径")
    parser.add_argument("--batch-size", type=int, default=None,
                        help="每次API调用合并的模板数")
    args = parser.parse_args()

    settings = load_settings(args.config)
    gbt_path = args.gbt or settings["gbt"]
    batch_size = args.batch_size or settings["batch_size"]
    retry_delay = settings["retry_delay"]
    model_configs = load_model_configs(args.config, settings["models"])

    module_dir = Path(__file__).resolve().parent / "module-list"
    xlsx_files = sorted(module_dir.glob("*.xlsx"))
    if not xlsx_files:
        print(f"未在 {module_dir} 下找到 xlsx 文件")
        return

    print(f"共 {len(xlsx_files)} 个文件待处理\n")
    for xf in xlsx_files:
        process_xlsx(xf, model_configs, gbt_path, batch_size, retry_delay)

    print(f"\n{'='*60}")
    print("全部处理完毕")


if __name__ == "__main__":
    main()
