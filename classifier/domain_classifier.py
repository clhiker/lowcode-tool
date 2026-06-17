#!/usr/bin/env python3
"""
基于模板描述，调用LLM生成：
  - 领域：GB/T 4754-2017 国民经济行业分类小类（1~3个）
  - 类型：从 type.txt 定义的模板类型中选取最匹配的 1 个

输出新的 xlsx，仅含三列：模板id、领域、类型，不修改原表格。

提示词定义在 prompts.py 中，可按需修改。

用法：
  python3 classifier/domain_classifier.py                            # 逐条
  python3 classifier/domain_classifier.py -n 5                       # 每5条合并一次调用
  python3 classifier/domain_classifier.py --xlsx xxx.xlsx -o out.xlsx
"""

import configparser
import re
import sys
import time
import argparse

import openpyxl
from openai import OpenAI

from prompts import SYSTEM_PROMPT, build_batch_prompt


def load_gbt_codes(gbt_path):
    codes = set()
    with open(gbt_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            m = re.match(r"(\d{4})\s+", line)
            if m:
                codes.add(m.group(1))
    print(f"已加载 {len(codes)} 个有效 GB/T 4754-2017 小类代码", file=sys.stderr)
    return codes


def extract_codes_from_text(text):
    codes = []
    for line in text.strip().split("\n"):
        line = line.strip()
        m = re.match(r"(\d{4})\s+\S", line)
        if m:
            codes.append(m.group(1))
    return codes


def extract_codes_from_domain_str(domain_str):
    return re.findall(r"\b(\d{4})\b", domain_str)


def validate_domains(domains_list, valid_codes):
    invalid_items = []
    for i, ds in enumerate(domains_list):
        codes = extract_codes_from_domain_str(ds)
        bad = [c for c in codes if c not in valid_codes]
        if bad:
            invalid_items.append((i, bad))
    return invalid_items


def build_correction_prompt(invalid_items, original_items):
    lines = ["注意：你上次返回中存在不存在的 GB/T 4754-2017 代码，请修正："]
    for idx, bad_codes in invalid_items:
        codes_str = ", ".join(bad_codes)
        lines.append(
            f"  <{idx + 1}> 中的 {codes_str} 不存在，请替换为真实存在的分类。"
        )
    lines.append(
        "只输出修正后的结果即可，格式与之前完全相同（每个模板一个块）。"
    )
    return "\n".join(lines)


def parse_batch_results(text, n):
    domains_list = [""] * n
    types_list = [""] * n
    current_idx = -1

    for line in text.strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"<(\d+)>\s*$", line)
        if m:
            current_idx = int(m.group(1)) - 1
            continue
        dm = re.match(r"领域[：:]\s*(.*)", line)
        if dm and 0 <= current_idx < n:
            domains_list[current_idx] = dm.group(1).strip()
            continue
        tm = re.match(r"类型[：:]\s*(.*)", line)
        if tm and 0 <= current_idx < n:
            types_list[current_idx] = tm.group(1).strip()

    for i in range(n):
        if not domains_list[i]:
            domains_list[i] = "通用（前端表单）"
        if not types_list[i]:
            types_list[i] = "组件模板"

    return domains_list, types_list


def call_api(client, model, messages, max_tokens=8192, timeout=180):
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=messages,
            temperature=0.1,
            max_tokens=max_tokens,
            timeout=timeout,
        )
        choice = resp.choices[0]
        content = choice.message.content
        if content is None:
            finish = choice.finish_reason
            print(f"  [警告] content为空, finish_reason={finish}", file=sys.stderr)
            return None
        return content.strip()
    except Exception as e:
        print(f"  [错误] API调用失败: {e}", file=sys.stderr)
        return None


def evaluate_formula(val, row_idx):
    s = str(val)
    m = re.search(r'CONCAT\("([^"]+)"\s*,\s*TEXT\s*\(\s*ROW\(\)\s*-\s*(\d+)\s*,\s*"([^"]+)"\s*\)\s*\)', s)
    if m:
        prefix = m.group(1)
        offset = int(m.group(2))
        fmt = m.group(3)
        num = row_idx - offset
        return f'{prefix}{num:0{len(fmt)}}'
    return s


def load_config(config_path, section):
    config = configparser.ConfigParser()
    config.read(config_path, encoding="utf-8")
    return {
        "base_url": config[section]["base_url"].rstrip("/"),
        "api_key": config[section]["api_key"],
        "model": config[section]["model"],
        "max_tokens": int(config[section].get("max_tokens", 4096)),
    }


def main():
    parser = argparse.ArgumentParser(
        description="从模板描述生成领域（GB/T 4754-2017）和类型标签"
    )
    parser.add_argument("--xlsx", default="classifier/51_FormVueLate_vue-模板结构.xlsx",
                        help="输入的xlsx文件，结果将直接写入该文件的新增列")
    parser.add_argument("--config", default="classifier/config.ini")
    parser.add_argument("--section", default="sjtu_minimax",
                        help="config.ini 中的配置节名称（默认 sjtu_minimax）")
    parser.add_argument("--gbt", default="classifier/gbt4754_2017_小类.txt",
                        help="GB/T 4754-2017 小类代码文件路径")
    parser.add_argument("-n", "--batch-size", type=int, default=20,
                        help="每次API调用合并的模板数（默认20）")
    parser.add_argument("--retry-until-valid", action="store_true",
                        help="无限重试直到所有代码校验通过（默认关闭）")
    parser.add_argument("--max-validation-retries", type=int, default=2,
                        help="校验失败后最大重试次数，仅在未启用 --retry-until-valid 时生效（默认2）")
    args = parser.parse_args()

    cfg = load_config(args.config, args.section)
    client = OpenAI(base_url=cfg["base_url"], api_key=cfg["api_key"])
    model = cfg["model"]

    valid_codes = load_gbt_codes(args.gbt)

    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb["汇总表"]

    col_id = 1
    col_func = 8
    col_desc = 11

    items = []
    for row_idx in range(2, ws.max_row + 1):
        tid_raw = ws.cell(row=row_idx, column=col_id).value
        desc = ws.cell(row=row_idx, column=col_desc).value
        if not tid_raw or not desc or not str(desc).strip():
            continue
        tid = evaluate_formula(tid_raw, row_idx)
        func = ws.cell(row=row_idx, column=col_func).value or ""
        items.append((row_idx, str(tid).strip(), str(desc).strip(), str(func).strip()))

    if not items:
        print("没有需要处理的模板。")
        return

    total = len(items)
    n = args.batch_size
    max_val_retries = float("inf") if args.retry_until_valid else args.max_validation_retries
    retry_label = "无限" if args.retry_until_valid else str(max_val_retries)
    print(f"共 {total} 个模板，batch-size={n}，校验重试={retry_label} 次\n")

    col_new1 = ws.max_column + 1
    col_new2 = col_new1 + 1
    ws.cell(row=1, column=col_new1).value = "新增列1"
    ws.cell(row=1, column=col_new2).value = "新增列2"

    all_domains = [""] * total
    all_types = [""] * total

    for start in range(0, total, n):
        batch = items[start:start + n]
        batch_size = len(batch)
        indices = list(range(start, start + batch_size))

        batch_items = [(d, f) for _, _, d, f in batch]
        mtokens = max(8192, batch_size * 2048)

        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": build_batch_prompt(batch_items)},
        ]

        success = False
        retry_history = ""
        val_attempt = 0
        while True:
            tag = f"[校验修正 {val_attempt}]" if val_attempt > 0 else ""
            print(f"[{start+1}-{start+batch_size}/{total}] 处理中{tag} ...", end=" ", flush=True)

            user_prompt = build_batch_prompt(batch_items)
            if retry_history:
                user_prompt += "\n\n" + retry_history

            text = call_api(client, model, [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ], max_tokens=mtokens)
            if not text:
                print("→ API调用失败")
                break

            domains_list, types_list = parse_batch_results(text, batch_size)
            invalid = validate_domains(domains_list, valid_codes)

            if not invalid:
                success = True
                print("→ 通过")
                for k, idx in enumerate(indices):
                    all_domains[idx] = domains_list[k]
                    all_types[idx] = types_list[k]
                break
            else:
                bad_summary = "; ".join(
                    f"<{i+1}>: {','.join(codes)}" for i, codes in invalid
                )
                print(f"→ 发现无效代码 ({bad_summary})")
                val_attempt += 1
                if val_attempt < max_val_retries or args.retry_until_valid:
                    retry_history = build_correction_prompt(invalid, batch_items)
                else:
                    for k, idx in enumerate(indices):
                        all_domains[idx] = domains_list[k]
                        all_types[idx] = types_list[k]
                    print(f"   重试耗尽，保留原始结果")
                    break

        if not success:
            for idx in indices:
                if not all_domains[idx]:
                    all_domains[idx] = "通用（前端表单）"
                if not all_types[idx]:
                    all_types[idx] = "组件模板"

        for k, (row_idx, _, _, _) in enumerate([items[idx] for idx in indices]):
            ws.cell(row=row_idx, column=col_new1).value = all_domains[indices[k]]
            ws.cell(row=row_idx, column=col_new2).value = all_types[indices[k]]
        wb.save(args.xlsx)
        time.sleep(0.5)

    print(f"\n完成！结果已写入 {args.xlsx}（新增列1=领域, 新增列2=类型）")


if __name__ == "__main__":
    main()
