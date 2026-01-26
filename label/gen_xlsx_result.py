import json
import re
import pandas as pd
import os
import argparse


def preprocess_result_md(file_path, file_dir, file_type):
    """
    预处理AI文档文件
    - 删除文件中的 ** 标记
    - 标准化功能描述（将换行替换为 ;）
    - 按数字顺序对结果进行排序
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 删除所有的 ** 标记
    content = content.replace('**', '')
    
    # 移除 ### 标记
    content = re.sub(r'###\s*', '', content)
    
    # 使用正则表达式解析内容
    # 匹配文件路径、功能标签和功能描述
    pattern = rf'({file_dir}/\d+\.{file_type})\s*\n\s*功能标签[：:]\s*(.*?)\s*\n+\s*功能描述[：:]\s*(.*?)(?=\n\s*{file_dir}/|$)'
    matches = re.findall(pattern, content, re.DOTALL)
    
    processed_data = []
    for match in matches:
        file_path = match[0].strip()
        function_tag = match[1].strip()
        function_desc = match[2].strip()
        
        # 标准化功能描述：将换行符替换为 ;
        function_desc = re.sub(r'\n+', '; ', function_desc)
        # 清理多余的空白字符
        function_desc = re.sub(r'\s+', ' ', function_desc)
        function_desc = function_desc.strip()
        
        processed_data.append({
            'file_path': file_path,
            'function_tag': function_tag,
            'function_desc': function_desc
        })
    
    # 按文件名中的数字进行排序，而不是按字符串排序
    processed_data.sort(key=lambda x: int(x['file_path'].split('/')[-1].split('.')[0]))
    
    return processed_data


def merge_with_excel(md_data, name, fixed_data, excel_path, output_path, seed_type=''):
    """
    将预处理后的md数据与Excel文件合并
    """
    # 读取Excel文件
    df = pd.read_excel(excel_path)
    
    # 创建一个字典，以便按文件名快速查找功能标签和描述
    # 将seed_AJSF/xxx.json格式转换为xxx.json格式进行匹配
    md_dict = {}
    for item in md_data:
        # 提取文件名部分，例如将 'seed_AJSF/375.json' 转换为 '375.json'
        full_file_path = item['file_path']
        file_name = full_file_path.split('/')[-1]  # 获取 '375.json'
        md_dict[file_name] = {
            'function_tag': item['function_tag'],
            'function_desc': item['function_desc']
        }
    fixed_dict = {}
    for item in fixed_data:
        # 提取文件名部分，例如将 'seed_AJSF/375.json' 转换为 '375.json'
        full_file_path = item['file_path']
        file_name = full_file_path.split('/')[-1]  # 获取 '375.json'
        fixed_dict[file_name] = {
            'function_tag': item['function_tag'],
            'function_desc': item['function_desc']
        }
    
    # 找出只存在于AI文档中的文件
    md_dict.update(fixed_dict)

    with open(f'final_md/{name}.json', 'w', encoding='utf-8') as f:
        json.dump(md_dict, f, ensure_ascii=False, indent=4)
    md_files = set(md_dict.keys())

    # 检查df的id列是否包含字符'.'，如果不包含，则给这一列的字符都增加一个seed_type
    def add_seed_type_if_no_dot(x):
        x_str = str(x)
        if '.' not in x_str:
            return f"{x_str}.{seed_type}"
        return x_str
    
    df['id'] = df['id'].apply(add_seed_type_if_no_dot)
    
    excel_files = set(df['id'].tolist())
    
    only_in_md = md_files - excel_files
    only_in_excel = excel_files - md_files
    
    print(f"只存在于AI文档中的文件数量: {len(only_in_md)}")
    if only_in_md:
        print("只存在于AI文档中的文件 (前10个):", list(only_in_md)[:10])
    
    print(f"只存在于原始文档中的文件数量: {len(only_in_excel)}")
    if only_in_excel:
        print("只存在于原始文档中的文件 (前10个):", list(only_in_excel)[:10])
        with open('miss.txt', 'w') as f:
            for file_name in only_in_excel:
                f.write(file_name + '\n')
    
    print(f"匹配的文件数量: {len(md_files & excel_files)}")
    
    # 添加两列到DataFrame
    # 使用id列进行匹配，id列包含如 '1.json', '2.json' 等值
    df['功能标签'] = df.apply(lambda row: md_dict.get(row['id'], {}).get('function_tag', ''), axis=1)
    df['功能描述'] = df.apply(lambda row: md_dict.get(row['id'], {}).get('function_desc', ''), axis=1)
    
    # 确保DataFrame按文件名中的数字排序
    df['numeric_id'] = df['id'].str.extract(r'(\d+)').astype(int)
    df = df.sort_values(by='numeric_id').drop('numeric_id', axis=1)
    
    # 保存合并后的Excel文件
    df.to_excel(output_path, index=False)
    print(f"合并完成！结果已保存到 {output_path}")


def create_template_structure(md_data, excel_path, standard_path, output_path, seed_type=''):
    """
    根据标准.xlsx创建模板结构.xlsx文件，并填入指定列的数据
    """
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment
    
    # 读取标准.xlsx文件作为模板结构，使用第2行作为列名
    standard_df = pd.read_excel(standard_path, header=1)
    
    # 创建一个字典，以便按文件名快速查找功能标签和描述
    md_dict = {}
    for item in md_data:
        full_file_path = item['file_path']
        file_name = full_file_path.split('/')[-1]
        md_dict[file_name] = {
            'function_tag': item['function_tag'],
            'function_desc': item['function_desc']
        }
    
    # 读取原始文档获取id, repo_full_name, original_html_url
    content_df = pd.read_excel(excel_path)
    
    # 检查content_df的id列是否包含字符'.'，如果不包含，则给这一列的字符都增加一个seed_type
    # 这是为了保持与merge_with_excel函数中相同的处理逻辑
    def add_seed_type_if_no_dot(x):
        x_str = str(x)
        if '.' not in x_str:
            return f"{x_str}.{seed_type}"
        return x_str
    
    if seed_type:
        content_df['id'] = content_df['id'].apply(add_seed_type_if_no_dot)
    
    # 确保内容 DataFrame 按文件名中的数字排序
    # 提取文件名中的数字部分用于排序
    content_df['numeric_id'] = content_df['id'].str.extract(r'(\d+)').astype(int)
    content_df = content_df.sort_values(by='numeric_id').drop('numeric_id', axis=1)
    
    # 创建新DataFrame，只包含需要填充的列
    new_data = []
    for index, row in content_df.iterrows():
        file_id = row['id']
        new_row = {
            '模板id': '',
            '模板名': '',
            '标签1. 语言和框架': '',  # 不填充
            '标签2. 文件后缀': '',  # 不填充
            '标签3. 用户库或Schema': '',  # 不填充
            '标签4. 原生模板/子模板/组合模板': '',  # 不填充
            '标签5. 领域': '',  # 不填充
            '标签6. 功能': md_dict.get(file_id, {}).get('function_tag', ''),  # 填充
            '提交者': '',  # 不填充
            '更新时间': '',  # 不填充
            '模板索引号': '',  # 不填充
            '模板描述': md_dict.get(file_id, {}).get('function_desc', ''),  # 填充
            '模板示例图片地址': '',  # 不填充
            '压缩包': '',  # 不填充
            '文件名': row['id'],  # 填充
            '代码地址': '',  # 不填充
            '文件所在仓库': row['repo_full_name'],  # 填充
            '文件来源': row['original_html_url'],  # 填充
            '时间戳': '',  # 不填充
            'token': '',  # 不填充
        }
        new_data.append(new_row)
    
    # 创建新的DataFrame，确保列顺序与标准文件一致
    result_df = pd.DataFrame(new_data)
    result_df = result_df.reindex(columns=standard_df.columns)
    
    # 直接复制标准文件，然后更新数据
    from openpyxl import load_workbook, Workbook
    standard_wb = load_workbook(standard_path)
    ws = standard_wb.active
    
    # 从第3行开始写入数据
    for idx, row in result_df.iterrows():
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=idx + 3, column=col_num)
            cell.value = value
    
    standard_wb.save(output_path)
    
    print(f"模板结构创建完成！结果已保存到 {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="根据AI生成的结果，对比github索引，生成和标准.xlsx一致的文件，在content目录下面保存着原始的xlsx文件")
    parser.add_argument("--name", type=str, help="处理的低代码模板名称")
    parser.add_argument("--file_type", type=str, help="种子模板的后缀")
    parser.add_argument("--fixed_path", default='fixed.md', type=str, help="修正的AI结果文件路径，可以为空")
    args = parser.parse_args()
    
    name = args.name
    # 处理AI文档文件
    seed_type = args.file_type
    ai_path = f'ai_result/{name}_result_corrected.md'
    fixed_path = args.fixed_path
    seed_path_name = f'seed_{name}'
    excel_path = f'content/{name}-content.xlsx'

    md_data = preprocess_result_md(ai_path, seed_path_name, seed_type)
    if os.path.exists(fixed_path):
        fixed_data = preprocess_result_md(fixed_path, seed_path_name, seed_type)
    else:
        fixed_data = []
 
    print("预处理完成，共处理了", len(md_data), "个文件的标签和描述")
    
    # 合并Excel文件
    output_path = 'merged.xlsx'
    merge_with_excel(md_data, name, fixed_data, excel_path, output_path, seed_type)
    
    # 创建模板结构文件
    standard_path = '标准.xlsx'
    template_output_path = f'{name}-模板结构.xlsx'
    
    create_template_structure(md_data, excel_path, standard_path, template_output_path, seed_type)
