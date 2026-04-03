import argparse
import json
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook
from tqdm import tqdm

CURRENT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = CURRENT_DIR.parent
if str(CURRENT_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_DIR))
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

try:
    from label.fill_xaml_template_by_ai import find_header_row_and_columns
    from label.fill_xaml_template_by_ai import get_pending_rows
    from label.fill_xaml_template_by_ai import normalize_header
except ImportError:
    from fill_xaml_template_by_ai import find_header_row_and_columns
    from fill_xaml_template_by_ai import get_pending_rows
    from fill_xaml_template_by_ai import normalize_header


ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_DIR = ROOT_DIR / "模板清单"
DEFAULT_OUTPUT_DIR = ROOT_DIR / "模板清单-new"
REPAIR_TOOL = Path(__file__).resolve().parent / "fill_xaml_template_by_ai.py"


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def strip_archive_suffix(name):
    value = clean_text(name)
    if not value:
        return ""

    suffixes = [".tar.gz", ".tgz", ".zip", ".tar", ".gz", ".rar", ".7z"]
    lowered = value.lower()
    changed = True
    while changed:
        changed = False
        for suffix in suffixes:
            if lowered.endswith(suffix):
                value = value[: -len(suffix)]
                lowered = value.lower()
                changed = True
                break
    return value


def find_col_index(ws, header_row, header_name):
    target = normalize_header(header_name)
    for col_idx in range(1, ws.max_column + 1):
        key = normalize_header(ws.cell(row=header_row, column=col_idx).value)
        if key == target:
            return col_idx
    return None


def find_first_data_row(ws, header_row, file_col):
    for row_idx in range(header_row + 1, ws.max_row + 1):
        file_name = clean_text(ws.cell(row=row_idx, column=file_col).value)
        if file_name:
            return row_idx
    return header_row + 1


def find_first_non_empty_in_column(ws, col_idx, start_row=1):
    if col_idx is None:
        return ""
    for row_idx in range(start_row, ws.max_row + 1):
        value = clean_text(ws.cell(row=row_idx, column=col_idx).value)
        if value:
            return value
    return ""


def resolve_source_dir(package_name):
    base = strip_archive_suffix(package_name)
    if not base:
        return None

    search_roots = [ROOT_DIR / "files", ROOT_DIR]
    for search_root in search_roots:
        exact = search_root / base
        if exact.is_dir():
            return exact

    for search_root in search_roots:
        candidates = [p for p in search_root.glob(f"{base}*") if p.is_dir()]
        if candidates:
            candidates.sort(key=lambda p: (len(p.name), p.name))
            return candidates[0]

    return None


def inspect_excel(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    header_row, columns = find_header_row_and_columns(ws)
    pending_count = len(get_pending_rows(ws, header_row, columns))

    syntax_col = find_col_index(ws, header_row, "标签1. 语言和框架")
    file_type_col = find_col_index(ws, header_row, "标签2. 文件后缀")
    package_col = find_col_index(ws, header_row, "压缩包")

    data_row = find_first_data_row(ws, header_row, columns["文件名"])

    syntax_type = clean_text(ws.cell(row=data_row, column=syntax_col).value)
    file_type = clean_text(ws.cell(row=data_row, column=file_type_col).value)
    package_name = clean_text(ws.cell(row=data_row, column=package_col).value)

    if not syntax_type:
        syntax_type = find_first_non_empty_in_column(ws, syntax_col, start_row=header_row + 1)
    if not file_type:
        file_type = find_first_non_empty_in_column(ws, file_type_col, start_row=header_row + 1)
    if not package_name:
        package_name = find_first_non_empty_in_column(ws, package_col, start_row=header_row + 1)

    wb.close()
    return {
        "pending": pending_count,
        "syntax_type": syntax_type,
        "file_type": file_type,
        "package_name": package_name,
    }


def run_repair_once(excel_path, source_dir, syntax_type, file_type, model, port):
    cmd = [
        sys.executable,
        str(REPAIR_TOOL),
        "--excel",
        str(excel_path),
        "--source_dir",
        str(source_dir),
        "--syntax_type",
        syntax_type,
        "--file_type",
        file_type,
        "--model",
        model,
        "--port",
        str(port),
    ]
    result = subprocess.run(cmd, cwd=str(ROOT_DIR), check=False)
    return result.returncode


def build_parser():
    parser = argparse.ArgumentParser(
        description="全量修复模板清单，循环调用 fill_xaml_template_by_ai.py 直到修复完成"
    )
    parser.add_argument("--input_dir", default=str(DEFAULT_INPUT_DIR), help="模板清单目录")
    parser.add_argument("--output_dir", default=str(DEFAULT_OUTPUT_DIR), help="修复输出目录")
    parser.add_argument("--model", default="qwen2.5:14b", help="Ollama 模型名")
    parser.add_argument("--port", default="11434", help="Ollama 服务端口")
    parser.add_argument(
        "--max_rounds",
        type=int,
        default=0,
        help="每个文件最多修复轮次，0 表示不设上限",
    )
    parser.add_argument(
        "--max_stall",
        type=int,
        default=10,
        help="连续无进展轮次数上限，0 表示不设上限",
    )
    return parser


def main():
    args = build_parser().parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)

    if not input_dir.exists():
        raise FileNotFoundError(f"输入目录不存在: {input_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)

    excel_files = sorted(input_dir.glob("*.xlsx"))
    if not excel_files:
        print(f"未找到 Excel 文件: {input_dir}")
        return

    prescan = {}
    total_pending = 0
    prescan_iter = tqdm(excel_files, desc="预扫描清单", unit="file")
    for excel in prescan_iter:
        info = inspect_excel(excel)
        prescan[excel.name] = info
        total_pending += info["pending"]
        prescan_iter.set_postfix_str(f"pending={total_pending}")

    print("=" * 60)
    print(f"共发现模板清单文件: {len(excel_files)}")
    print(f"修复前待补全总行数: {total_pending}")
    print("=" * 60)

    summary = []
    total_repaired = 0
    unresolved_total = 0

    file_iter = tqdm(excel_files, desc="文件修复进度", unit="file")
    for idx, src_excel in enumerate(file_iter, start=1):
        info = prescan[src_excel.name]
        before_pending = info["pending"]
        syntax_type = info["syntax_type"]
        file_type = info["file_type"]
        package_name = info["package_name"]
        file_iter.set_postfix_str(src_excel.name[:40])

        dst_excel = output_dir / src_excel.name
        shutil.copy2(src_excel, dst_excel)

        if before_pending == 0:
            print(f"[{idx}/{len(excel_files)}] {src_excel.name}: 无需修复，已拷贝")
            summary.append(
                {
                    "file": src_excel.name,
                    "status": "copied",
                    "before_pending": 0,
                    "after_pending": 0,
                    "repaired_rows": 0,
                    "rounds": 0,
                    "syntax_type": syntax_type,
                    "file_type": file_type,
                    "package_name": package_name,
                    "source_dir": "",
                }
            )
            continue

        source_dir = resolve_source_dir(package_name)
        if source_dir is None:
            print(
                f"[{idx}/{len(excel_files)}] {src_excel.name}: 未找到源目录，跳过修复 "
                f"(压缩包字段: {package_name})"
            )
            unresolved_total += before_pending
            summary.append(
                {
                    "file": src_excel.name,
                    "status": "source_dir_missing",
                    "before_pending": before_pending,
                    "after_pending": before_pending,
                    "repaired_rows": 0,
                    "rounds": 0,
                    "syntax_type": syntax_type,
                    "file_type": file_type,
                    "package_name": package_name,
                    "source_dir": "",
                }
            )
            continue

        rounds = 0
        stall_count = 0
        prev_pending = before_pending

        print(
            f"[{idx}/{len(excel_files)}] {src_excel.name}: 开始修复，"
            f"待补全 {before_pending} 行，source_dir={source_dir.name}"
        )

        rounds_bar = tqdm(
            total=args.max_rounds if args.max_rounds > 0 else None,
            desc=f"轮次 {src_excel.name[:24]}",
            unit="round",
            leave=False,
        )

        while prev_pending > 0:
            if args.max_rounds > 0 and rounds >= args.max_rounds:
                break

            rounds += 1
            rounds_bar.update(1)
            rc = run_repair_once(
                dst_excel,
                source_dir,
                syntax_type,
                file_type,
                args.model,
                args.port,
            )

            current_pending = inspect_excel(dst_excel)["pending"]
            rounds_bar.set_postfix_str(f"remain={current_pending}")
            print(
                f"  - 第 {rounds} 轮: exit={rc}, 剩余待补全 {current_pending} 行"
            )

            if current_pending == 0:
                prev_pending = 0
                break

            if current_pending < prev_pending:
                stall_count = 0
            else:
                stall_count += 1

            prev_pending = current_pending

            if args.max_stall > 0 and stall_count >= args.max_stall:
                print(
                    f"  - 连续无进展达到 {args.max_stall} 轮，停止当前文件继续重试"
                )
                break

        rounds_bar.close()

        after_pending = inspect_excel(dst_excel)["pending"]
        repaired_rows = before_pending - after_pending
        total_repaired += max(repaired_rows, 0)
        unresolved_total += after_pending

        status = "success" if after_pending == 0 else "partial"
        summary.append(
            {
                "file": src_excel.name,
                "status": status,
                "before_pending": before_pending,
                "after_pending": after_pending,
                "repaired_rows": repaired_rows,
                "rounds": rounds,
                "syntax_type": syntax_type,
                "file_type": file_type,
                "package_name": package_name,
                "source_dir": str(source_dir),
            }
        )

    report = {
        "input_dir": str(input_dir),
        "output_dir": str(output_dir),
        "files_total": len(excel_files),
        "pending_before_total": total_pending,
        "repaired_total": total_repaired,
        "pending_after_total": unresolved_total,
        "results": summary,
    }

    report_path = output_dir / "repair_summary.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print("=" * 60)
    print(f"输出目录: {output_dir}")
    print(f"汇总文件: {report_path}")
    print(f"总修复行数: {total_repaired}")
    print(f"剩余未修复行数: {unresolved_total}")
    print("每个文件修复行数:")
    for item in summary:
        print(
            f"- {item['file']}: repaired={item['repaired_rows']}, "
            f"before={item['before_pending']}, after={item['after_pending']}, status={item['status']}"
        )


if __name__ == "__main__":
    main()