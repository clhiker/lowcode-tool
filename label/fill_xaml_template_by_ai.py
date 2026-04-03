import argparse
import os
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from tqdm import tqdm


def normalize_header(value):
    if value is None:
        return ""
    text = str(value)
    text = text.replace(" ", "").replace("\u3000", "")
    return text.strip()


def find_header_row_and_columns(ws):
    required = {
        "标签6.功能": None,
        "模板描述": None,
        "文件名": None,
    }

    for row_idx in range(1, min(50, ws.max_row) + 1):
        found = dict(required)
        for col_idx in range(1, ws.max_column + 1):
            key = normalize_header(ws.cell(row=row_idx, column=col_idx).value)
            if key in found and found[key] is None:
                found[key] = col_idx
        if all(v is not None for v in found.values()):
            return row_idx, found

    raise ValueError("未找到表头行，必须包含列：标签6. 功能、模板描述、文件名")


def is_empty_cell(value):
    if value is None:
        return True
    return str(value).strip() == ""


def clean_ai_field(text):
    if text is None:
        return ""
    v = str(text).strip()
    v = re.sub(r"^[`'\"*\-:：\s]+", "", v)
    v = re.sub(r"\s+", " ", v).strip()
    return v


def is_meaningful_text(text):
    if not text:
        return False
    # 至少包含一个中文、字母或数字，避免把 ':' 这类符号当正文
    return re.search(r"[\u4e00-\u9fffA-Za-z0-9]", text) is not None


def is_bad_label_text(label_text):
    if not label_text:
        return True
    if not is_meaningful_text(label_text):
        return True
    # 模型经常以“当然...”开头解释性话术，这不是有效功能标签
    return label_text.startswith("当然")


def is_bad_desc_text(desc_text):
    if not desc_text:
        return True
    if not is_meaningful_text(desc_text):
        return True
    # 对话式回复不是模板描述正文
    return desc_text.startswith("请问您")


def needs_fill(label_val, desc_val):
    label_text = clean_ai_field(label_val)
    desc_text = clean_ai_field(desc_val)
    label_bad = is_bad_label_text(label_text)
    desc_bad = is_bad_desc_text(desc_text)
    return label_bad or desc_bad


def extract_label_desc(ai_text):
    cleaned = ai_text.replace("**", "").strip()

    label = ""
    desc = ""

    # 优先按“单行键值”解析
    label_match = re.search(r"^\s*功能标签\s*[：:]\s*(.*?)\s*$", cleaned, re.MULTILINE)
    desc_match = re.search(r"^\s*功能描述\s*[：:]\s*(.*?)\s*$", cleaned, re.MULTILINE)

    if label_match:
        label = clean_ai_field(label_match.group(1))
    if desc_match:
        desc = clean_ai_field(desc_match.group(1))

    # 兼容“下一行才是正文”的输出格式
    lines = [line.strip() for line in cleaned.splitlines() if line.strip()]
    if lines and (not is_meaningful_text(label) or not is_meaningful_text(desc)):
        for i, line in enumerate(lines):
            if not is_meaningful_text(label) and re.match(r"^功能标签\s*[：:]?\s*$", line):
                if i + 1 < len(lines):
                    label = clean_ai_field(lines[i + 1])
            if not is_meaningful_text(desc) and re.match(r"^功能描述\s*[：:]?\s*$", line):
                if i + 1 < len(lines):
                    desc = clean_ai_field(lines[i + 1])

    # 最后兜底：取第一行/后续行，但必须是有意义文本
    if lines:
        if not is_meaningful_text(label):
            candidate = re.sub(r"^功能标签\s*[：:]?\s*", "", lines[0]).strip()
            label = clean_ai_field(candidate)
        if not is_meaningful_text(desc):
            rest = " ".join(lines[1:]) if len(lines) > 1 else ""
            rest = re.sub(r"^功能描述\s*[：:]?\s*", "", rest).strip()
            desc = clean_ai_field(rest)

    if not is_meaningful_text(label):
        label = ""
    if not is_meaningful_text(desc):
        desc = ""

    label = label[:20]
    desc = re.sub(r"\s+", " ", desc).strip()
    return label, desc


def build_prompt(source_text, syntax_type, file_type):
    text_snippet = source_text[:3000]
    return [
        {
            "role": "system",
            "content": f"你是一个前端开发专家，对于{syntax_type}语法有很深的了解。你的回答不要包含 ** 符号。",
        },
        {
            "role": "user",
            "content": (
                f"请读取下面的{file_type}文件，并输出两行内容：\n"
                "功能标签：不超过8个汉字。\n"
                "功能描述：60-100个中文字符，不分段，不包含子标签和markdown符号。\n"
                "文件内容如下：\n"
                f"{text_snippet}\n"
                "请中文回答。"
            ),
        },
    ]


def backup_excel(excel_path):
    src = Path(excel_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = src.with_name(f"{src.stem}.backup_{ts}{src.suffix}")
    shutil.copy2(src, backup_path)
    return backup_path


def get_pending_rows(ws, header_row, columns):
    label_col = columns["标签6.功能"]
    desc_col = columns["模板描述"]
    file_col = columns["文件名"]

    pending = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        file_name = ws.cell(row=row_idx, column=file_col).value
        if file_name is None or str(file_name).strip() == "":
            continue

        label_val = ws.cell(row=row_idx, column=label_col).value
        desc_val = ws.cell(row=row_idx, column=desc_col).value

        if needs_fill(label_val, desc_val):
            pending.append(row_idx)
    return pending


def count_incomplete(ws, header_row, columns):
    label_col = columns["标签6.功能"]
    desc_col = columns["模板描述"]
    file_col = columns["文件名"]

    incomplete = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        file_name = ws.cell(row=row_idx, column=file_col).value
        if file_name is None or str(file_name).strip() == "":
            continue
        label_val = ws.cell(row=row_idx, column=label_col).value
        desc_val = ws.cell(row=row_idx, column=desc_col).value
        if needs_fill(label_val, desc_val):
            incomplete += 1
    return incomplete


def main():
    parser = argparse.ArgumentParser(
        description="补全模板结构.xlsx中 标签6. 功能/模板描述 空值（调用本地 Ollama）"
    )
    parser.add_argument(
        "--excel",
        default="xaml-模板结构.xlsx",
        help="待处理 Excel 路径（默认：xaml-模板结构.xlsx）",
    )
    parser.add_argument("--source_dir", default="xaml", help="源文件目录（默认：xaml）")
    parser.add_argument(
        "--syntax_type",
        default="xaml",
        help="语法种类，用于提示词（默认：xaml）",
    )
    parser.add_argument(
        "--file_type",
        default="xaml",
        help="文件类型，用于提示词与缺失文件兜底后缀（默认：xaml）",
    )
    parser.add_argument(
        "--model",
        default="qwen2.5:14b",
        help="Ollama 模型名（默认：qwen2.5:14b）",
    )
    parser.add_argument(
        "--port",
        default="11434",
        help="Ollama 服务端口（默认：11434）",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="最多处理多少行，0 表示处理全部",
    )
    parser.add_argument(
        "--check_only",
        action="store_true",
        help="只检查空值，不执行 AI 生成",
    )

    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    source_dir = Path(args.source_dir)
    if not source_dir.exists():
        raise FileNotFoundError(f"源文件目录不存在: {source_dir}")

    wb = load_workbook(excel_path)
    ws = wb.active

    header_row, columns = find_header_row_and_columns(ws)
    all_pending_rows = get_pending_rows(ws, header_row, columns)
    all_pending_count = len(all_pending_rows)

    if args.limit > 0:
        pending_rows = all_pending_rows[: args.limit]
    else:
        pending_rows = all_pending_rows

    print(f"检测到待补全总行数: {all_pending_count}")
    print(f"本次计划处理行数: {len(pending_rows)}")

    if args.check_only:
        return

    if not pending_rows:
        print("无需处理，文件已完整。")
        return

    backup_path = backup_excel(excel_path)
    print(f"已备份原始文件: {backup_path}")

    os.environ["OLLAMA_HOST"] = f"127.0.0.1:{args.port}"
    try:
        from ollama import chat
    except ImportError as exc:
        raise ImportError("未安装 ollama Python 包，请先执行: pip install ollama") from exc

    label_col = columns["标签6.功能"]
    desc_col = columns["模板描述"]
    file_col = columns["文件名"]

    total = len(pending_rows)
    done = 0
    failed = 0

    row_iter = tqdm(pending_rows, total=total, desc="行修复进度", unit="row")
    for idx, row_idx in enumerate(row_iter, start=1):
        file_name = str(ws.cell(row=row_idx, column=file_col).value).strip()
        source_path = source_dir / file_name

        row_iter.set_postfix_str(file_name[:40])

        if not source_path.exists() and "." not in file_name:
            source_path = source_dir / f"{file_name}.{args.file_type.lstrip('.')}"

        if not source_path.exists():
            failed += 1
            print(f"[{idx}/{total}] 缺少源文件，跳过: {source_path}")
            continue

        with open(source_path, "r", encoding="utf-8", errors="ignore") as f:
            source_text = f.read()

        try:
            response = chat(
                model=args.model,
                messages=build_prompt(source_text, args.syntax_type, args.file_type),
            )
            ai_text = ""
            if hasattr(response, "message") and hasattr(response.message, "content"):
                ai_text = response.message.content
            elif isinstance(response, dict):
                ai_text = str(response.get("message", {}).get("content", ""))
            label, desc = extract_label_desc(ai_text)
            if not label or not desc:
                raise ValueError("AI 返回内容无法解析")
        except Exception as exc:
            failed += 1
            print(f"[{idx}/{total}] AI 生成失败 {file_name}: {exc}")
            continue

        old_label = ws.cell(row=row_idx, column=label_col).value
        old_desc = ws.cell(row=row_idx, column=desc_col).value

        old_label_text = clean_ai_field(old_label)
        old_desc_text = clean_ai_field(old_desc)

        if is_bad_label_text(old_label_text):
            ws.cell(row=row_idx, column=label_col).value = label
        if is_bad_desc_text(old_desc_text):
            ws.cell(row=row_idx, column=desc_col).value = desc

        done += 1
        if done % 20 == 0 or idx == total:
            wb.save(excel_path)
        print(f"[{idx}/{total}] 已补全: {file_name}")

    wb.save(excel_path)

    wb2 = load_workbook(excel_path)
    ws2 = wb2.active
    _, columns2 = find_header_row_and_columns(ws2)
    remaining = count_incomplete(ws2, header_row, columns2)

    print("=" * 60)
    print(f"处理完成: 成功 {done} 行, 失败 {failed} 行")
    print(f"剩余未补全行数: {remaining}")
    if args.limit > 0 and all_pending_count > len(pending_rows):
        print(
            f"说明: 因设置了 --limit {args.limit}，仅处理了前 {len(pending_rows)} 行待补全记录。"
        )
    if remaining == 0:
        print("校验结果: 已完全补完。")
    else:
        print("校验结果: 尚未完全补完，请检查失败日志后重试。")


if __name__ == "__main__":
    main()
